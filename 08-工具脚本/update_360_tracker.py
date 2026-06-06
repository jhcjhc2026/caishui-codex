# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


INPUT = r"C:\codex_tmp\phone_lead_tracker_fix_input.xlsx"
OUTPUT = r"C:\codex_tmp\phone_lead_tracker_360.xlsx"


def main():
    wb = load_workbook(INPUT)
    ws = wb.worksheets[0]

    headers = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    industry_col = headers.get("行业判断") or 5
    script_col = headers.get("首次开场角度") or headers.get("360元代账开场话术") or 8

    scripts = {
        "餐饮食品": "新开餐饮公司前期最怕漏报税、成本票不规范。我们现在有新注册小规模公司基础代账 360 元/月起，包含每月记账、纳税申报和税务提醒。先帮您确认一下：现在账务报税这块安排好了吗？",
        "建筑工程": "建筑/工程类公司后面容易涉及合同、开票、成本票和用工问题。我们针对新注册小规模公司有基础代账 360 元/月起，先把记账报税和税务节点规范起来。您现在是自己处理，还是已经找代账了？",
        "商贸零售": "商贸公司开票和进项成本会比较关键。我们这边新注册小规模公司基础代账 360 元/月起，包含记账、报税和开票税务提醒。想问下您公司现在每月报税这块有人负责了吗？",
        "咨询服务": "咨询/管理类公司前期通常票量不大，但也要按月申报、规范成本费用。我们有新注册小规模公司基础代账 360 元/月起，适合前期轻运营公司。您现在代账报税安排好了吗？",
        "科技互联网": "科技类公司前期除了报税，也要注意研发费用、成本票和后续政策申报基础。我们新注册小规模公司基础代账 360 元/月起，先把每月记账申报规范起来。您现在财税这块是自己做还是外包了？",
        "文化传媒": "文化传媒公司经常会涉及服务费、劳务费、开票和成本归集。我们新注册小规模公司基础代账 360 元/月起，包含记账、报税和基础财税提醒。您现在有固定会计或代账公司了吗？",
        "实业制造": "实业/制造类公司后面成本、库存、设备票据会更重要。前期如果票量不大，可以先用 360 元/月起的基础代账把申报规范起来。您现在账务这块安排好了吗？",
        "人力资源": "人力资源/劳务类公司通常要关注合同、开票、个税和用工相关风险。我们基础代账 360 元/月起，适合新公司前期先把申报和提醒做规范。您现在财税有人负责了吗？",
        "汽车交通": "汽车/物流类公司后续容易涉及油费、维修、运输票据和开票问题。我们新注册小规模公司基础代账 360 元/月起，先把记账报税和票据提醒做好。您现在代账安排好了吗？",
        "健康医药": "健康/医药类公司对票据和合规要求会更敏感。我们新注册小规模公司基础代账 360 元/月起，包含每月申报和基础税务提醒。您现在账务报税这块有人在处理吗？",
        "教育培训": "教育培训类公司前期即使暂时没收入，也要注意零申报、年报和票据规范。我们新注册小规模公司基础代账 360 元/月起，适合前期先低成本规范运营。您现在报税安排好了吗？",
        "房地产": "房产/物业类公司后续开票、合同和成本归集会比较关键。我们新注册小规模公司基础代账 360 元/月起，先把每月申报和税务提醒做好。您现在财税这块安排了吗？",
        "其他/待确认": "您好，我这边是上海本地财税服务的。看到贵公司是新注册不久，我们现在有新注册小规模公司基础代账 360 元/月起，包含每月记账、纳税申报和税务提醒。想先确认一下，您现在记账报税这块安排好了吗？",
    }

    for r in range(2, ws.max_row + 1):
        industry = ws.cell(r, industry_col).value or "其他/待确认"
        ws.cell(r, script_col).value = scripts.get(industry, scripts["其他/待确认"])
        ws.cell(r, script_col).alignment = Alignment(vertical="top", wrap_text=True)

    ws.cell(1, script_col).value = "360元代账开场话术"
    ws.column_dimensions[get_column_letter(script_col)].width = 58
    ws.cell(1, script_col).fill = PatternFill("solid", fgColor="C65911")
    ws.cell(1, script_col).font = Font(color="FFFFFF", bold=True)
    ws.auto_filter.ref = ws.dimensions

    if len(wb.worksheets) >= 3:
        sws = wb.worksheets[2]
        sws.delete_rows(1, sws.max_row)
    else:
        sws = wb.create_sheet("电话话术")
    sws.title = "电话话术"

    rows = [
        ["场景", "话术", "目标", "记录建议"],
        ["首次开场", scripts["其他/待确认"], "确认是否决策人，抛出 360 元/月基础代账卖点。", "是否本人/是否已有代账/是否愿意加微信"],
        ["对方感兴趣", "360 元/月是基础小规模代账价，适合新公司前期票量不多、需要正常报税和避免异常的情况。我们会先根据行业、开票量、员工情况确认是否适用。我加您微信，把服务项和新公司注意事项发您看一下可以吗？", "争取加微信并确认需求。", "票量/员工/行业/微信号/预计开业时间"],
        ["已有代账", "没关系，很多新公司刚开始都会先找一家做基础申报。我们这个 360 元/月主要是给您做个对比：除了报税，还会提醒开票、成本票和年报等节点。您先留个微信，以后对价格或服务不满意时也有个备选。", "用“对比/备选”降低防备。", "现有代账价格/不满点/到期时间"],
        ["问 360 包含什么", "360 元/月一般包含记账、纳税申报、财税节点提醒和日常基础咨询。具体要看票量、银行流水、员工社保和行业情况。我先加您微信，发一份清单，您对照着看更清楚。", "说明边界，避免过度承诺。", "关注项/是否要清单"],
        ["嫌便宜不放心", "这个顾虑很正常。360 元不是把服务做少，而是针对新公司前期票量少、业务还没复杂的阶段。后面票量上来或有员工、出口、高新等情况，再按实际需求调整。前期先用合适成本把申报做规范。", "强调阶段适配。", "质量顾虑/复杂度/是否要案例"],
        ["嫌贵或再看看", "可以理解。新公司最关键的是不要因为漏报、错报、年报遗漏导致异常。360 元/月就是把基础财税安排好。您可以先加我微信，我把服务项和注意事项发您，对比后再决定。", "转成避免异常成本。", "预算/下次跟进时间"],
        ["暂时没经营", "暂时没经营也没问题，但注册后税务申报和工商年报节点还是要关注。如果是零申报、票量少，360 元/月这种基础服务就比较适合。我先加您微信，发一份新公司财税时间表给您。", "用零申报和异常风险承接。", "是否开业/预计开业时间"],
        ["对方很忙", "好的，我不耽误您太久。就一句话：我们现在有新注册公司基础代账 360 元/月起，适合前期票量不多的小规模公司。我加您微信发资料，您有空看一眼就行，可以吗？", "快速争取加微信。", "回拨时间/是否加微"],
        ["拒绝/不需要", "好的，那我不多打扰。新公司前期最容易忽略的是每月申报、发票风险和工商年报。后面如果您需要对比 360 元/月的基础代账方案，我再给您发资料。祝您生意顺利。", "礼貌收口，留二次触达可能。", "拒绝原因/是否可二次跟进"],
    ]
    for row in rows:
        sws.append(row)

    for c in sws[1]:
        c.fill = PatternFill("solid", fgColor="1F4E78")
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in sws.iter_rows(min_row=2):
        row[0].font = Font(bold=True, color="1F4E78")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for i, width in enumerate([20, 90, 30, 38], 1):
        sws.column_dimensions[get_column_letter(i)].width = width
    for r in range(2, sws.max_row + 1):
        sws.row_dimensions[r].height = 75
    sws.freeze_panes = "A2"
    sws.auto_filter.ref = sws.dimensions

    wb.save(OUTPUT)
    print(OUTPUT)
    print("updated_rows=", ws.max_row - 1)
    print("main_header=", ws.cell(1, script_col).value)
    print("main_sample=", ws.cell(2, script_col).value[:80])
    print("script_rows=", sws.max_row)


if __name__ == "__main__":
    main()
